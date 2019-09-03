# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.core.mail import send_mail
from django.conf import settings
from finapp.forms import RegForm, InvestReg
from finapp.models import Fintab, UserTrack
from openpyxl import load_workbook
import requests
import json
import datetime
# Create your views here.
def index(request):
	#del request.session['name']
	if request.session.has_key('name'):
		return render(request,'index.html',{'name':request.session['name']})
	else:
		return render(request,'index.html',{})
	
def catagory(request):
	#del request.session['fname']
	return render(request,'catagory.html',{})
def finadv(request):
	wb=load_workbook('./finapp/ChoiceData.xlsx')
	compsheet=wb['Company Name Picklist']
	citysheet=wb['City State Picklist']
	compname=[]
	ciname=[]
	for i in range(2,(compsheet.max_row + 1)):
		compname.append(compsheet.cell(row=i,column=1).value)
	for i in range(2,(citysheet.max_row + 1)):
		ciname.append(citysheet.cell(row=i,column=1).value)
	return render(request,'register.html',{'compname':compname,'ciname':ciname})
def investPro(request):
	wb=load_workbook('./finapp/ChoiceData.xlsx')
	compsheet=wb['Company Name Picklist']
	citysheet=wb['City State Picklist']
	compname=[]
	ciname=[]
	for i in range(2,(compsheet.max_row + 1)):
		compname.append(compsheet.cell(row=i,column=1).value)
	for i in range(2,(citysheet.max_row + 1)):
		ciname.append(citysheet.cell(row=i,column=1).value)
	return render(request,'invest.html',{'compname':compname,'ciname':ciname})
def asstMan(request):
	wb=load_workbook('./finapp/ChoiceData.xlsx')
	compsheet=wb['Company Name Picklist']
	citysheet=wb['City State Picklist']
	compname=[]
	ciname=[]
	for i in range(2,(compsheet.max_row + 1)):
		compname.append(compsheet.cell(row=i,column=1).value)
	for i in range(2,(citysheet.max_row + 1)):
		ciname.append(citysheet.cell(row=i,column=1).value)
	return render(request,'asst.html',{'compname':compname,'ciname':ciname})
	
def register(request):
	if request.method == "POST":
		fname=request.POST.get('fname')
		request.session['fname']=fname
		request.session['nname']=request.POST.get('nname')
		request.session['sname']=request.POST.get('sname')
		request.session['cname']=request.POST.get('cname')
		request.session['cityname']=request.POST.get('cityname')
		request.session['email']=request.POST.get('email')
		request.session['phone']=request.POST.get('phone')
		request.session['uname']=request.POST.get('uname')
		request.session['psw']=request.POST.get('psw')
		request.session['pswrpt']=request.POST.get('pswrpt')
		if request.session['psw'] ==request.session['pswrpt']:
			fullname=fname+" "+request.session['sname']
			stname=request.session['cityname']
			stname=stname.split(',')[1]
			stname=stname.strip()
			link="https://api.brokercheck.finra.org/individual?hl=true&includePrevious=true&json.wrf=angular.callbacks._2&nrows=12&query="+fullname+"&r=25&sort=score+desc&state="+stname+"&wt=json"
			#request.session['result']=JsonResponse('https://api.brokercheck.finra.org/individual?hl=true&includePrevious=true&json.wrf=angular.callbacks._2&nrows=12&query=Max+Sparshatt&r=25&sort=score+desc&state=NY&wt=json',safe=False)
			response=requests.get(link)
			data=response.content
			data=data.decode('utf-8')
			data=data.replace('/**/angular.callbacks._2(','')
			data=data.replace(');','')
			res=json.loads(data)
			if res['hits']['total'] > 0:
				firstname=res['hits']['hits'][0]['_source']['ind_firstname']
				midname=res['hits']['hits'][0]['_source']['ind_middlename']
				lastname=res['hits']['hits'][0]['_source']['ind_lastname']
				crd=res['hits']['hits'][0]['_source']['ind_source_id']
				disclosure=res['hits']['hits'][0]['_source']['ind_bc_disclosure_fl']
				firm_count=res['hits']['hits'][0]['_source']['ind_employments_count']
				firm_crd=res['hits']['hits'][0]['_source']['ind_current_employments'][0]['firm_id']
				firm_name=res['hits']['hits'][0]['_source']['ind_current_employments'][0]['firm_name']
				city=res['hits']['hits'][0]['_source']['ind_current_employments'][0]['branch_city']
				state=res['hits']['hits'][0]['_source']['ind_current_employments'][0]['branch_state']
				bzip=res['hits']['hits'][0]['_source']['ind_current_employments'][0]['branch_zip']
				lcount=res['hits']['hits'][0]['_source']['ind_approved_finra_registration_count']
				request.session['crd']=crd
				#userdata=Fintab(firstname=fname,midname=request.session['nname'],lastname=request.session['sname'],companyname=request.session['cname'],
				#cityname=request.session['cityname'],email=request.session['email'],phone=request.session['phone'],uname=request.session['uname'],
				#psw=request.session['psw'],pswrpt=request.session['pswrpt'])
				#userdata.save()
				return render(request,'confirm.html',{"fname":firstname,"midname":midname,"lastname":lastname,"crd":crd,"firm_name":firm_name,"firm_crd":firm_crd,"city":city,"state":state,"bzip":bzip,"firm_count":firm_count,"disclosure":disclosure,"lcount":lcount})

			else:
				return HttpResponse('<h2>User Does not exist</h2> <a href="finadv">Try Again or Go Back</a>')		
		else:
			return HttpResponse("<h1>Password is not matching</h1><a href='finadv'> Go To Registration Page</a>")		
	else:
		#Myregform=RegForm()
		return HttpResponse('<h2>Page Refreshing Restricted Try Again</h2> <a href="finadv"> Go To Registration Page</a>')

def acverify(request):
	ifname=request.session['fname']
	inname=request.session['nname']
	isname=request.session['sname']
	icname=request.session['cname']
	icityname=request.session['cityname']
	iemail=request.session['email']
	iphone=request.session['phone']
	iuname=request.session['uname']
	ipsw=request.session['psw']
	ipswrpt=request.session['pswrpt']
	crd=request.session['crd']
	x=datetime.datetime.now()
	member=x.strftime("%B")+" "+str(x.year)
	try:
		check_user=Fintab.objects.get(uname=iuname)
		return HttpResponse("<h1>User name already Exists ! Choose another one.</h1><a href='finadv'>Go Back</a>")
	except:
		ivpro=Fintab(firstname=ifname,midname=inname,lastname=isname,companyname=icname,cityname=icityname,email=iemail,phone=iphone,uname=iuname,psw=ipsw,pswrpt=ipswrpt,crd=crd,protype="FinAdv",member=member)
		ivpro.save()
		check=Fintab.objects.get(email=iemail)
		title="Greetings From Finra"
		subject="Hello "+check.firstname+", You have registered Successfully with Finra. Click this link to get verified and login your account "+"https://finmanageapp.herokuapp.com/finapp/check?id="+str(check.id)
		sender="robert.wheeler.mtp@gmail.com"
		send_mail(title,subject,sender,[iemail])
		return render(request,'acverify.html',{})

def investForm(request):
	if request.method =="POST":
		ifname=request.POST.get('ifname')
		inname=request.POST.get('inname')
		isname=request.POST.get('isname')
		icname=request.POST.get('icname')
		icityname=request.POST.get('icityname')
		iuname=request.POST.get('iuname')
		iemail=request.POST.get('iemail')
		iphone=request.POST.get('iphone')
		ipsw=request.POST.get('ipsw')
		ipswrpt=request.POST.get('ipswrpt')
		x=datetime.datetime.now()
		member=x.strftime("%B")+" "+str(x.year)
		try:
			check_user=Fintab.objects.get(uname=iuname)
			return HttpResponse("<h1>User name already Exists</h1><a href='investPro'>Go Back</a>")
		except:
			if ipsw == ipswrpt:
				ivpro=Fintab(firstname=ifname,midname=inname,lastname=isname,companyname=icname,cityname=icityname,email=iemail,phone=iphone,uname=iuname,psw=ipsw,pswrpt=ipswrpt,crd="",protype="InvestPro",member=member)
				ivpro.save()
				check=Fintab.objects.get(email=iemail)
				title="Greetings From Finra"
				subject="Hello "+check.firstname+", You have registered Successfully with Finra. Click this link to get verified and login your account "+"https://finmanageapp.herokuapp.com/finapp/check?id="+str(check.id)
				sender="robert.wheeler.mtp@gmail.com"
				send_mail(title,subject,sender,[iemail])
				return render(request,'acverify.html',{})
			else:
				return HttpResponse("<h1>Password is not matching</h1><a href='investPro'>Go to Registration Page</a>")
	else:
		return HttpResponse('<h1>Enter Data And then Register</h1><a href="investPro">Go to Registration Page</a>')
			

def asstForm(request):
	if request.method =="POST":
		afname=request.POST.get('afname')
		anname=request.POST.get('anname')
		asname=request.POST.get('asname')
		acname=request.POST.get('acname')
		acityname=request.POST.get('acityname')
		aemail=request.POST.get('aemail')
		aphone=request.POST.get('aphone')
		auname=request.POST.get('auname')
		apsw=request.POST.get('apsw')
		apswrpt=request.POST.get('apswrpt')
		x=datetime.datetime.now()
		member=x.strftime('%B')+" "+str(x.year)
		try:
			check_user=Fintab.objects.get(uname=auname)
			return HttpResponse("<h1>User name already Exists</h1><a href='investPro'>Go Back</a>")
		except:
			if apsw == apswrpt:
				ivpro=Fintab(firstname=afname,midname=anname,lastname=asname,companyname=acname,cityname=acityname,email=aemail,phone=aphone,uname=auname,psw=apsw,pswrpt=apswrpt,crd="",protype="AsstMan",member=member)
				ivpro.save()
				check=Fintab.objects.get(email=aemail)
				title="Greetings From Finra"
				subject="Hello "+check.firstname+", You have registered Successfully with Finra. Click this link to get verified and login your account "+"https://finmanageapp.herokuapp.com/finapp/check?id="+str(check.id)
				sender="robert.wheeler.mtp@gmail.com"
				send_mail(title,subject,sender,[aemail])
				return render(request,'acverify.html',{})
			else:
				return HttpResponse("<h1 >Password is not matching</h1><a href='asstMan'>Go to Registration Page</a>")
	else:
		return HttpResponse('<h1>Enter Data And then Register</h1><a href="asstMan">Go to Registration Page</a>')

def check(request):
	id=request.GET.get('id')
	try:
		user_name=Fintab.objects.get(id=id)
		request.session['id']=user_name.id
		request.session['name']=user_name.firstname
	except:
		return HttpResponse('<h1 style="margin-left:25%; margin-top: 10%">You are not registered ! Please Register First</h1><a href="catagory" style="margin-left: 40%">Goto Registration Page..</a>')
	return render(request,'index.html',{'name':user_name.firstname})

def help(request):
	if request.session.has_key('name'):
		return render(request,'help.html',{'name':request.session['name']})
	else:
		return render(request,'help.html',{})

def profile(request):
	if request.session.has_key('name'):
		user_info=Fintab.objects.get(id=request.session['id'])
		firstname=user_info.firstname
		lastname=user_info.lastname
		midname=user_info.midname
		uname=user_info.uname
		psw=user_info.psw
		member_since=user_info.member
		compname=user_info.companyname
		address=user_info.cityname
		phone=user_info.phone
		email=user_info.email
		crd=user_info.crd
		fullname=firstname+" "+lastname
		if crd =="":
			return render(request,'profile.html',{'name':request.session['name'],'fname':fullname,'uname':uname,'midname':midname,'uname':uname,'psw':psw,'phone':phone,'email':email,'crd':crd,'cname':compname,'address':address,'member_since':member_since})
		else:
			return render(request,'profile.html',{'name':request.session['name'],'fname':fullname,'uname':uname,'midname':midname,'uname':uname,'psw':psw,'phone':phone,'email':email,'crd':crd,'cname':compname,'address':address,'member_since':member_since})
	else:
		return render(request,'index.html',{})

def validate_user(request):
	username=request.GET.get('uname')
	data={'is_taken':Fintab.objects.filter(uname=username).exists()}
	return JsonResponse(data)

def editname(request):
	if request.session.has_key('name'):
		return render(request,'editname.html',{'name':request.session['name']})
	else:
		return render(request,'index.html',{})

def updatename(request):
	if request.method == "POST":
		midname=request.POST.get('nname')
		if request.session.has_key('id'):
			data=Fintab.objects.filter(id=request.session['id']).update(midname=midname)
			track=UserTrack(userid=request.session['id'],act="Nick Name Updated")
			track.save()
			return HttpResponse('<h1 style="margin-left:25%; margin-top: 10%"> Preferred Name Updated ! Please Check Your Profile Page</h1><a href="profile" style="margin-left: 40%">Goto Profile Page..</a>')
	else:
		return HttpResponse("<h1> Page Refreshing Prohibited </h1> <a href='editname'>Go Back</a>")		

def edituser(request):
	if request.session.has_key('name'):
		return render(request,'edituser.html',{'name':request.session['name']})
	else:
		return render(request,'index.html',{})

def updateuser(request):
	if request.method == "POST":
		oldp=request.POST.get('oldp')
		dbpsw=Fintab.objects.get(id=request.session['id'])
		if oldp == dbpsw.psw:
			newp=request.POST.get('newp')
			newrp=request.POST.get('newrp')
			if newp == newrp:
					if request.session.has_key('id'):
						data=Fintab.objects.filter(id=request.session['id']).update(psw=newp)
						track=UserTrack(userid=request.session['id'],act="Password Updated")
						track.save()
						return HttpResponse('<h1 style="margin-left:25%; margin-top: 10%"> Password Updated ! Please Check Your Profile Page</h1><a href="profile" style="margin-left: 40%">Goto Profile Page..</a>')
			else:
				return HttpResponse("<h1> Password is not matching</h1><a href='edituser'>Try Once Again</a>")
		else:
			return HttpResponse("<h1> Password is not matching. Please Enter Your Old Password Correctly</h1><a href='edituser'>Try Once Again</a>")
	else:
		return HttpResponse("<h1> Page Refreshing Prohibited </h1> <a href='edituser'>Go Back</a>")				


def editphone(request):
	if request.session.has_key('name'):
		return render(request,'editphone.html',{'name':request.session['name']})
	else:
		return render(request,'index.html',{})

def updatephone(request):
	if request.method == "POST":
		phone=request.POST.get('phone')
		if request.session.has_key('id'):
			data=Fintab.objects.filter(id=request.session['id']).update(phone=phone)
			track=UserTrack(userid=request.session['id'],act="Phone no Updated")
			track.save()
			return HttpResponse('<h1 style="margin-left:25%; margin-top: 10%"> Phone Number Updated ! Please Check Your Profile Page</h1><a href="profile" style="margin-left: 40%">Goto Profile Page..</a>')
	else:
		return HttpResponse("<h1> Page Refreshing Prohibited </h1> <a href='editphone'>Go Back</a>")


def editmail(request):
	if request.session.has_key('name'):
		return render(request,'editmail.html',{'name':request.session['name']})	
	else:
		return render(request,'index.html',{})

def updatemail(request):
	if request.method == "POST":
		email=request.POST.get('email')
		if request.session.has_key('id'):
			data=Fintab.objects.filter(id=request.session['id']).update(email=email)
			track=UserTrack(userid=request.session['id'],act="Email ID Updated")
			track.save()
			return HttpResponse('<h1 style="margin-left:25%; margin-top: 10%"> Email ID Updated ! Please Check Your Profile Page</h1><a href="profile" style="margin-left: 40%">Goto Profile Page..</a>')
	else:
		return HttpResponse("<h1> Page Refreshing Prohibited </h1> <a href='editmail'>Go Back</a>")

def check_send_mail(request):
	send_mail("Testing", "Testing Send mail","robert.wheeler.mtp@gmail.com",["itssara6@gmail.com"])
	return HttpResponse('<h1>Mail send</h1>')





